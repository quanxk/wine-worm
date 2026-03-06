import requests
from bs4 import BeautifulSoup
import lxml
import csv
import pandas as pd
import xlwt
import xlrd
import re
from openpyxl.workbook import Workbook


def get_2020_wine():
        file_name = 'farr-2020.xls'
        print('您选择导入的文件是： ', file_name)
        print('请耐心等待列表导入')
        target_wookbook = xlrd.open_workbook(file_name)
        target_table = target_wookbook.sheet_by_index(0)
        workbook = xlwt.Workbook(encoding='utf-8')
        writebook = xlwt.Workbook(r'farr.xlsx')
        cell_overwrite_ok = True
        sheet = writebook.add_sheet('2018farr', cell_overwrite_ok)

        row = 1

        sheet.write(row, 3, 'FARR')
        sheet.write(row, 4, 'NM barrel')
        sheet.write(row, 5, 'NM Bottle')
        sheet.write(row, 6, 'WA Barrel')
        sheet.write(row, 7, 'WA Bottle')
        sheet.write(row, 8, 'JS Barrel')
        sheet.write(row, 9, 'JS Bottle')
        sheet.write(row, 10, 'MS')
        sheet.write(row, 11, 'JLMW')
        sheet.write(row, 12, 'JA')
        for i in range(0, target_table.nrows):
            target = target_table.cell(i, 0).value
            print(target)

            r1 = requests.get(target, timeout=20)
            soup1 = BeautifulSoup(r1.text, 'lxml')
            div_title = soup1.find('title')
            print(div_title.text)
            div_sub = soup1.find_all('div',{'class', True})
            print(div_sub)
            # sheet.write(row, 2, wine_info_link_full)
            row = row + 1
            for i in div_sub:

                sheet.write(row, 1, div_title.text)
                div_farr = i.find('span', {'class','notranslate'})
                div_name_farr = i.find('i')
                try:
                    div_name_farr_con = div_name_farr.text
                    div_farr_point = div_farr.text
                    div_name_sum= div_name_farr_con
                    if div_name_sum == 'Farr Vintners, Farr Tasting, April 2021':
                        print(div_name_farr_con)
                        print(div_farr_point)
                        sheet.write(row, 3, div_farr_point)
                    elif div_name_sum == 'Neal Martin, vinous.com, May 2021':
                        print(div_name_farr_con)
                        print(div_farr_point)
                        sheet.write(row, 4, div_farr_point)
                    elif div_name_sum == 'Neal Martin, vinous.com, March 2021':
                        print(div_name_farr_con)
                        print(div_farr_point)
                        sheet.write(row, 5, div_farr_point)
                    elif div_name_sum == 'Lisa Perrotti-Brown MW, Wine Advocate, May 2021':
                        print(div_name_farr_con)
                        print(div_farr_point)
                        sheet.write(row, 6, div_farr_point)
                    elif div_name_sum == 'Lisa Perrotti-Brown MW, Wine Advocate (End of Mar), March 2021':
                        print(div_name_farr_con)
                        print(div_farr_point)
                        sheet.write(row, 7, div_farr_point)
                    elif div_name_sum == 'James Suckling, JamesSuckling.com, April 2021':
                        print(div_name_farr_con)
                        print(div_farr_point)
                        sheet.write(row, 8, div_farr_point)
                    elif div_name_sum == 'James Suckling, JamesSuckling.com, March 2021':
                        print(div_name_farr_con)
                        print(div_farr_point)
                        sheet.write(row, 9, div_farr_point)
                    elif div_name_sum == 'Michael Schuster, The World of Fine Wine, May 2021':
                        print(div_name_farr_con)
                        print(div_farr_point)
                        sheet.write(row, 10, div_farr_point)
                    elif div_name_sum == 'James Lawther MW, JancisRobinson.com, April 2021':
                        print(div_name_farr_con)
                        print(div_farr_point)
                        sheet.write(row, 11, div_farr_point)
                    elif div_name_sum == 'Jane Anson, Decanter.com, May 2021':
                        print(div_name_farr_con)
                        print(div_farr_point)
                        sheet.write(row, 12, div_farr_point)

                except AttributeError:
                    pass


        writebook.save('farr-2020-barrel.xlsx')





def main():

    get_2020_wine()



if __name__ == '__main__':
        main()